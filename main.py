import streamlit as st
import asyncio
import json
import io
from datetime import datetime
from typing import Dict, Any, Optional
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import validators

from analysis_engine import AIEntityAnalyzer
from data_models import AnalysisReport

# Page configuration
st.set_page_config(
    page_title="AI SEO Entity Analyzer",
    page_icon="🔍",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for better styling
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #1f77b4;
        text-align: center;
        margin-bottom: 1rem;
    }
    .success-box {
        background-color: #d4edda;
        border: 1px solid #c3e6cb;
        color: #155724;
        padding: 1rem;
        border-radius: 0.5rem;
        margin: 1rem 0;
    }
    .error-box {
        background-color: #f8d7da;
        border: 1px solid #f5c6cb;
        color: #721c24;
        padding: 1rem;
        border-radius: 0.5rem;
        margin: 1rem 0;
    }
    .warning-box {
        background-color: #fff3cd;
        border: 1px solid #ffeaa7;
        color: #856404;
        padding: 1rem;
        border-radius: 0.5rem;
        margin: 1rem 0;
    }
    .metric-container {
        background-color: #f8f9fa;
        padding: 1rem;
        border-radius: 0.5rem;
        border-left: 4px solid #1f77b4;
    }
</style>
""", unsafe_allow_html=True)

class SEOAnalyzerApp:
    def __init__(self):
        self.init_session_state()
    
    def init_session_state(self):
        """Initialize session state variables."""
        if 'analysis_report' not in st.session_state:
            st.session_state.analysis_report = None
        if 'analysis_history' not in st.session_state:
            st.session_state.analysis_history = []
        if 'form_submitted' not in st.session_state:
            st.session_state.form_submitted = False
        if 'analysis_running' not in st.session_state:
            st.session_state.analysis_running = False
    
    def validate_api_key(self) -> bool:
        """Validate OpenAI API key."""
        try:
            api_key = st.secrets.get("OPENAI_API_KEY")
            if not api_key:
                st.error("❌ OpenAI API key not found in secrets")
                st.info("Please add your OpenAI API key to the secrets.toml file")
                return False
            return True
        except Exception as e:
            st.error(f"❌ Error accessing API key: {str(e)}")
            return False
    
    def render_header(self):
        """Render the main header."""
        st.markdown('<div class="main-header">🔍 AI SEO Entity Analyzer</div>', unsafe_allow_html=True)
        st.markdown("**Discover SEO opportunities by analyzing competitor content with AI**")
        st.markdown("---")
    
    def render_sidebar(self):
        """Render the sidebar with navigation."""
        with st.sidebar:
            st.markdown("## 🎯 Navigation")
            
            # Analysis history
            if st.session_state.analysis_history:
                st.markdown("### 📋 Recent Analyses")
                for i, analysis in enumerate(st.session_state.analysis_history[-3:]):
                    if st.button(f"📊 Analysis {i+1}", key=f"load_{i}"):
                        st.session_state.analysis_report = analysis
                        st.session_state.form_submitted = False
                        st.rerun()
            
            st.markdown("---")
            
            # Quick stats
            if st.session_state.analysis_report:
                report = st.session_state.analysis_report
                st.markdown("### 📈 Current Analysis")
                st.metric("Opportunities Found", report.total_opportunities)
                st.metric("Priority Items", report.priority_opportunities)
                st.metric("Competitors", len(report.competitor_analyses))
            
            # Reset button
            if st.session_state.analysis_report or st.session_state.form_submitted:
                st.markdown("---")
                if st.button("🔄 Start New Analysis", type="secondary"):
                    st.session_state.analysis_report = None
                    st.session_state.form_submitted = False
                    st.session_state.analysis_running = False
                    st.rerun()
            
            st.markdown("---")
            st.markdown("### ℹ️ About")
            st.info(
                "This tool uses AI to analyze your content against competitors "
                "and find SEO opportunities. Simply enter your URL and competitor URLs to get started."
            )
    
    def render_input_form(self):
        """Render the main input form with improved handling."""
        st.markdown("## 🌐 Enter URLs for Analysis")
        
        # Show warning if analysis is running
        if st.session_state.analysis_running:
            st.warning("⏳ Analysis in progress... Please wait for completion.")
            return
        
        # Initialize form state
        if 'client_url' not in st.session_state:
            st.session_state.client_url = ""
        if 'competitor_urls' not in st.session_state:
            st.session_state.competitor_urls = ["", "", ""]
        
        # Client URL input
        st.markdown("### 🏠 Your Website")
        client_url = st.text_input(
            "Your Website URL",
            value=st.session_state.client_url,
            placeholder="https://your-website.com",
            help="Enter the URL of the page you want to analyze",
            key="client_url_input"
        )
        
        # Update session state
        st.session_state.client_url = client_url
        
        # Competitor URLs
        st.markdown("### 🏢 Competitor URLs")
        st.info("💡 **Tip**: Enter at least 2 competitor URLs for meaningful analysis. All URLs will be validated before analysis starts.")
        
        competitor_urls = []
        for i in range(3):
            url = st.text_input(
                f"Competitor {i+1} URL",
                value=st.session_state.competitor_urls[i] if i < len(st.session_state.competitor_urls) else "",
                placeholder=f"https://competitor{i+1}.com",
                key=f"comp_url_{i}"
            )
            if url.strip():
                competitor_urls.append(url.strip())
            
            # Update session state
            if i < len(st.session_state.competitor_urls):
                st.session_state.competitor_urls[i] = url
        
        # Analysis options
        st.markdown("### ⚙️ Analysis Options")
        col1, col2 = st.columns(2)
        
        with col1:
            detailed_analysis = st.checkbox("Detailed Analysis Mode", value=False, help="More thorough analysis but takes longer")
        
        with col2:
            max_opportunities = st.selectbox("Max Opportunities to Find", [5, 10, 15, 20], index=1)
        
        st.markdown("---")
        
        # Validation and submit section
        col1, col2 = st.columns([3, 1])
        
        with col1:
            # Real-time validation display
            validation_messages = []
            
            if not client_url:
                validation_messages.append("❌ Client URL is required")
            elif not validators.url(client_url):
                validation_messages.append("❌ Client URL format is invalid")
            else:
                validation_messages.append("✅ Client URL is valid")
            
            if len(competitor_urls) < 2:
                validation_messages.append("❌ At least 2 competitor URLs are required")
            else:
                valid_competitors = [url for url in competitor_urls if validators.url(url)]
                invalid_competitors = [url for url in competitor_urls if not validators.url(url)]
                
                validation_messages.append(f"✅ {len(valid_competitors)} valid competitor URLs")
                if invalid_competitors:
                    validation_messages.append(f"❌ {len(invalid_competitors)} invalid competitor URLs")
            
            # Display validation status
            for msg in validation_messages:
                if "❌" in msg:
                    st.error(msg)
                else:
                    st.success(msg)
        
        with col2:
            # Submit button
            can_submit = (
                client_url and 
                validators.url(client_url) and 
                len([url for url in competitor_urls if validators.url(url)]) >= 2 and
                not st.session_state.analysis_running
            )
            
            if st.button(
                "🚀 Start Analysis", 
                disabled=not can_submit,
                use_container_width=True,
                type="primary"
            ):
                # Validate one more time before submission
                valid_competitor_urls = [url for url in competitor_urls if validators.url(url)]
                
                if client_url and validators.url(client_url) and len(valid_competitor_urls) >= 2:
                    st.session_state.form_submitted = True
                    st.session_state.analysis_running = True
                    self.run_analysis(client_url, valid_competitor_urls, detailed_analysis, max_opportunities)
                else:
                    st.error("❌ Please fix validation errors before submitting")
    
    def run_analysis(self, client_url: str, competitor_urls: list, detailed_analysis: bool = False, max_opportunities: int = 10):
        """Run the analysis with comprehensive progress tracking and error handling."""
        
        # Create analysis container
        analysis_container = st.container()
        
        with analysis_container:
            st.markdown("## 🔄 Analysis in Progress")
            
            # Progress tracking
            progress_bar = st.progress(0)
            status_text = st.empty()
            details_container = st.container()
            
            # Error tracking
            errors = []
            warnings = []
            
            def update_progress(message: str, progress: float = None):
                status_text.text(f"🔍 {message}")
                if progress is not None:
                    progress_bar.progress(min(100, int(progress * 100)))
            
            def log_error(error_msg: str):
                errors.append(error_msg)
                with details_container:
                    st.error(f"❌ {error_msg}")
            
            def log_warning(warning_msg: str):
                warnings.append(warning_msg)
                with details_container:
                    st.warning(f"⚠️ {warning_msg}")
            
            def log_success(success_msg: str):
                with details_container:
                    st.success(f"✅ {success_msg}")
        
        try:
            # Get API key
            api_key = st.secrets.get("OPENAI_API_KEY")
            if not api_key:
                log_error("OpenAI API key not found in secrets")
                st.session_state.analysis_running = False
                return
            
            log_success("API key validated successfully")
            update_progress("Initializing analysis engine...", 0.1)
            
            # Run analysis with detailed tracking
            async def run_async_analysis():
                async with AIEntityAnalyzer(api_key) as analyzer:
                    try:
                        # Enhanced progress callback
                        def enhanced_progress_callback(message: str):
                            if "scraping" in message.lower():
                                update_progress(message, 0.2)
                            elif "analyzing client" in message.lower():
                                update_progress(message, 0.4)
                            elif "analyzing competitor" in message.lower():
                                update_progress(message, 0.6)
                            elif "identifying" in message.lower():
                                update_progress(message, 0.8)
                            elif "generating" in message.lower():
                                update_progress(message, 0.9)
                            elif "complete" in message.lower():
                                update_progress(message, 1.0)
                            else:
                                update_progress(message)
                        
                        # Start analysis
                        update_progress("Starting comprehensive analysis...", 0.1)
                        log_success(f"Analyzing client: {client_url}")
                        log_success(f"Analyzing {len(competitor_urls)} competitors")
                        
                        report = await analyzer.run_complete_analysis(
                            client_url, competitor_urls, enhanced_progress_callback
                        )
                        
                        return report
                        
                    except Exception as e:
                        log_error(f"Analysis engine error: {str(e)}")
                        return None
            
            # Execute async analysis
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            try:
                report = loop.run_until_complete(run_async_analysis())
            finally:
                loop.close()
            
            # Process results
            if report:
                # Analysis successful
                progress_bar.progress(100)
                status_text.success("✅ Analysis completed successfully!")
                
                # Store results
                st.session_state.analysis_report = report
                st.session_state.analysis_history.append(report)
                st.session_state.form_submitted = True
                st.session_state.analysis_running = False
                
                # Show summary
                with details_container:
                    st.markdown("### 📊 Analysis Summary")
                    col1, col2, col3, col4 = st.columns(4)
                    
                    with col1:
                        st.metric("Client Words", report.client_analysis.word_count)
                    
                    with col2:
                        st.metric("Competitors Analyzed", len(report.competitor_analyses))
                    
                    with col3:
                        st.metric("Opportunities Found", report.total_opportunities)
                    
                    with col4:
                        st.metric("Priority Items", report.priority_opportunities)
                    
                    # Analysis quality check
                    if report.total_opportunities == 0:
                        st.warning("⚠️ **No opportunities found!** This could indicate:")
                        st.markdown("""
                        - Your content is already well-optimized compared to competitors
                        - Competitors have similar content strategies
                        - Analysis parameters may need adjustment
                        - Consider analyzing different competitor pages or industries
                        """)
                    
                    if len(report.competitor_analyses) < len(competitor_urls):
                        failed_count = len(competitor_urls) - len(report.competitor_analyses)
                        st.warning(f"⚠️ **{failed_count} competitor(s) failed to analyze** - they may be blocking scraping or have accessibility issues")
                
                # Auto-scroll to results
                st.rerun()
                
            else:
                # Analysis failed
                log_error("Analysis failed to complete successfully")
                st.session_state.analysis_running = False
                
                with details_container:
                    st.markdown("### 🔧 Troubleshooting Tips")
                    st.info("""
                    **Common issues and solutions:**
                    - **URL Access**: Ensure all URLs are publicly accessible
                    - **Content Length**: Pages with very little text content may not analyze well
                    - **Rate Limits**: If you see rate limit errors, wait a few minutes and try again
                    - **API Quota**: Check your OpenAI account for available credits
                    """)
                    
                    if st.button("🔄 Retry Analysis"):
                        st.session_state.analysis_running = False
                        st.rerun()
                        
        except Exception as e:
            log_error(f"Unexpected error during analysis: {str(e)}")
            st.session_state.analysis_running = False
            
            with details_container:
                st.markdown("### 🆘 Error Details")
                st.code(str(e))
                
                if st.button("🔄 Reset and Try Again"):
                    st.session_state.analysis_running = False
                    st.session_state.form_submitted = False
                    st.rerun()
    
    def render_results(self):
        """Render analysis results with enhanced debugging info."""
        if not st.session_state.analysis_report:
            return
        
        report = st.session_state.analysis_report
        
        st.markdown("## 📊 Analysis Results")
        
        # Enhanced summary metrics
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Total Opportunities", report.total_opportunities)
        with col2:
            st.metric("Priority Items", report.priority_opportunities)
        with col3:
            st.metric("Competitors Analyzed", len(report.competitor_analyses))
        with col4:
            st.metric("Client Word Count", report.client_analysis.word_count)
        
        # Debug information
        with st.expander("🔍 Debug Information"):
            st.markdown("**Analysis Metadata:**")
            st.json({
                "analysis_id": report.analysis_id,
                "timestamp": report.timestamp.isoformat(),
                "client_url": report.client_url,
                "competitor_urls_submitted": len(report.competitor_urls),
                "competitor_analyses_completed": len(report.competitor_analyses),
                "client_entities_found": len(report.client_analysis.entities),
                "total_competitor_entities": sum(len(comp.entities) for comp in report.competitor_analyses),
                "opportunities_found": len(report.missing_entities),
                "recommendations_generated": len(report.recommendations)
            })
            
            if len(report.competitor_analyses) != len(report.competitor_urls):
                st.warning("⚠️ Some competitors failed to analyze - check URLs for accessibility")
        
        # Tabs for different views
        tab1, tab2, tab3, tab4 = st.tabs(["🎯 Opportunities", "📋 Recommendations", "📈 Visualizations", "📁 Export"])
        
        with tab1:
            self.render_opportunities_tab(report)
        
        with tab2:
            self.render_recommendations_tab(report)
        
        with tab3:
            self.render_visualizations_tab(report)
        
        with tab4:
            self.render_export_tab(report)
    
    def render_opportunities_tab(self, report: AnalysisReport):
        """Render opportunities tab with better insights."""
        st.markdown("### 🎯 SEO Opportunities Found")
        
        if not report.missing_entities:
            st.info("🎉 **No opportunities found!** This could mean:")
            st.markdown("""
            - ✅ Your content is already well-optimized compared to these competitors
            - ✅ You're covering the same key topics and entities as your competitors
            - 🔄 Try analyzing different competitors in your industry
            - 🔄 Consider analyzing competitor pages that rank higher for your target keywords
            """)
            return
        
        st.markdown(f"Found **{len(report.missing_entities)} optimization opportunities** by analyzing competitor content:")
        
        for i, opportunity in enumerate(report.missing_entities, 1):
            with st.expander(f"{i}. {opportunity.name} (Relevance: {opportunity.relevance_score:.2f})"):
                col1, col2 = st.columns(2)
                
                with col1:
                    st.markdown(f"**Entity Type:** {opportunity.type}")
                    st.markdown(f"**Relevance Score:** {opportunity.relevance_score:.2f}")
                    st.markdown(f"**Avg. Competitor Salience:** {opportunity.avg_competitor_salience:.2f}")
                
                with col2:
                    st.markdown(f"**Found in {len(opportunity.found_in_competitors)} competitors:**")
                    for comp_url in opportunity.found_in_competitors:
                        st.markdown(f"• {comp_url[:50]}{'...' if len(comp_url) > 50 else ''}")
                
                st.markdown(f"**💡 Why this is an opportunity:** {opportunity.reasoning}")
    
    def render_recommendations_tab(self, report: AnalysisReport):
        """Render recommendations tab."""
        st.markdown("### 🚀 AI-Powered Recommendations")
        
        if not report.recommendations:
            st.info("No specific recommendations available. This can happen when:")
            st.markdown("""
            - No clear opportunities were identified in the analysis
            - The content analysis didn't find sufficient differences between your page and competitors
            - The AI model encountered issues generating specific recommendations
            """)
            return
        
        st.markdown(f"**{len(report.recommendations)} actionable recommendations** to improve your content:")
        
        for i, rec in enumerate(report.recommendations, 1):
            st.markdown(f"### {i}. {rec.entity_name}")
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown(f"**Target Section:** {rec.section}")
                st.markdown(f"**Recommendation:** {rec.recommendation}")
            
            with col2:
                st.markdown(f"**Expected SEO Impact:** {rec.seo_impact}")
            
            st.markdown("**Example Integration:**")
            st.code(rec.example_text, language="text")
            
            st.markdown("---")
    
    def render_visualizations_tab(self, report: AnalysisReport):
        """Render visualizations tab."""
        st.markdown("### 📈 Data Visualizations")
        
        # Opportunity relevance chart
        if report.missing_entities:
            st.subheader("Opportunity Relevance Scores")
            
            entity_names = [entity.name for entity in report.missing_entities]
            relevance_scores = [entity.relevance_score for entity in report.missing_entities]
            
            fig = px.bar(
                x=entity_names,
                y=relevance_scores,
                title="SEO Opportunities by Relevance Score",
                labels={"x": "Entities", "y": "Relevance Score"},
                color=relevance_scores,
                color_continuous_scale="viridis"
            )
            fig.update_layout(height=500)
            st.plotly_chart(fig, use_container_width=True)
        
        # Sentiment comparison
        st.subheader("Sentiment Analysis Comparison")
        
        sources = ["Your Website"] + [f"Competitor {i+1}" for i in range(len(report.competitor_analyses))]
        sentiments = [report.client_analysis.overall_sentiment] + [comp.overall_sentiment for comp in report.competitor_analyses]
        
        fig = px.bar(
            x=sources,
            y=sentiments,
            title="Content Sentiment Comparison",
            labels={"x": "Source", "y": "Sentiment Score"},
            color=sentiments,
            color_continuous_scale="RdYlBu"
        )
        fig.update_layout(height=400)
        st.plotly_chart(fig, use_container_width=True)
        
        # Word count comparison
        st.subheader("Content Length Comparison")
        
        word_counts = [report.client_analysis.word_count] + [comp.word_count for comp in report.competitor_analyses]
        
        fig = px.bar(
            x=sources,
            y=word_counts,
            title="Word Count Comparison",
            labels={"x": "Source", "y": "Word Count"},
            color=word_counts,
            color_continuous_scale="blues"
        )
        fig.update_layout(height=400)
        st.plotly_chart(fig, use_container_width=True)
    
    def render_export_tab(self, report: AnalysisReport):
        """Render export options tab."""
        st.markdown("### 📁 Export Analysis Results")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.markdown("#### 📊 Excel Report")
            if st.button("Generate Excel Report"):
                excel_data = self.create_excel_report(report)
                st.download_button(
                    label="Download Excel Report",
                    data=excel_data,
                    file_name=f"seo_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        
        with col2:
            st.markdown("#### 📄 JSON Data")
            if st.button("Generate JSON Export"):
                json_data = report.to_json()
                st.download_button(
                    label="Download JSON Data",
                    data=json_data,
                    file_name=f"seo_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                    mime="application/json"
                )
        
        with col3:
            st.markdown("#### 📝 Summary Report")
            if st.button("Generate Summary"):
                summary = self.create_summary_report(report)
                st.download_button(
                    label="Download Summary",
                    data=summary,
                    file_name=f"seo_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
                    mime="text/plain"
                )
    
    def create_excel_report(self, report: AnalysisReport) -> bytes:
        """Create Excel report from analysis results."""
        output = io.BytesIO()
        workbook = Workbook()
        
        # Remove default sheet
        workbook.remove(workbook.active)
        
        # Header styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Summary sheet
        summary_sheet = workbook.create_sheet("Summary")
        summary_data = [
            ["Metric", "Value"],
            ["Analysis Date", report.timestamp.strftime("%Y-%m-%d %H:%M:%S")],
            ["Client URL", report.client_url],
            ["Total Opportunities", report.total_opportunities],
            ["Priority Opportunities", report.priority_opportunities],
            ["Competitors Analyzed", len(report.competitor_analyses)]
        ]
        
        for row in summary_data:
            summary_sheet.append(row)
        
        # Style headers
        for cell in summary_sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        # Opportunities sheet
        opportunities_sheet = workbook.create_sheet("Opportunities")
        opp_headers = ["Entity", "Type", "Relevance Score", "Avg Salience", "Found in URLs", "Reasoning"]
        opportunities_sheet.append(opp_headers)
        
        for cell in opportunities_sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        for opp in report.missing_entities:
            opportunities_sheet.append([
                opp.name,
                opp.type,
                opp.relevance_score,
                opp.avg_competitor_salience,
                "; ".join(opp.found_in_competitors),
                opp.reasoning
            ])
        
        # Recommendations sheet
        rec_sheet = workbook.create_sheet("Recommendations")
        rec_headers = ["Entity", "Section", "Recommendation", "Example", "SEO Impact"]
        rec_sheet.append(rec_headers)
        
        for cell in rec_sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        for rec in report.recommendations:
            rec_sheet.append([
                rec.entity_name,
                rec.section,
                rec.recommendation,
                rec.example_text,
                rec.seo_impact
            ])
        
        workbook.save(output)
        output.seek(0)
        return output.getvalue()
    
    def create_summary_report(self, report: AnalysisReport) -> str:
        """Create a text summary report."""
        summary = f"""
SEO ENTITY ANALYSIS REPORT
=========================

Analysis Date: {report.timestamp.strftime("%Y-%m-%d %H:%M:%S")}
Client URL: {report.client_url}

SUMMARY METRICS
--------------
Total Opportunities Found: {report.total_opportunities}
Priority Opportunities: {report.priority_opportunities}
Competitors Analyzed: {len(report.competitor_analyses)}

TOP OPPORTUNITIES
----------------
"""
        
        for i, opp in enumerate(report.missing_entities[:5], 1):
            summary += f"{i}. {opp.name} (Score: {opp.relevance_score:.2f})\n"
            summary += f"   Type: {opp.type}\n"
            summary += f"   Reasoning: {opp.reasoning}\n\n"
        
        summary += "KEY RECOMMENDATIONS\n"
        summary += "-------------------\n"
        
        for i, rec in enumerate(report.recommendations[:3], 1):
            summary += f"{i}. {rec.entity_name}\n"
            summary += f"   Section: {rec.section}\n"
            summary += f"   Recommendation: {rec.recommendation}\n"
            summary += f"   Example: {rec.example_text}\n\n"
        
        return summary
    
    def run(self):
        """Run the main application."""
        self.render_header()
        
        if not self.validate_api_key():
            return
        
        self.render_sidebar()
        
        if st.session_state.analysis_report:
            self.render_results()
        else:
            self.render_input_form()

def main():
    """Main application entry point."""
    app = SEOAnalyzerApp()
    app.run()

if __name__ == "__main__":
    main()
