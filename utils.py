import io
import json
import tempfile
from typing import Dict, Any, List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from urllib.parse import urlparse
import validators
import streamlit as st
from datetime import datetime
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px

def get_competitor_name(url: str, used_domains: Dict[str, int]) -> str:
    """Generate a unique competitor name based on domain."""
    try:
        parsed_url = urlparse(url)
        domain = parsed_url.netloc.replace('www.', '')
        
        if not domain:
            return f"unknown_domain_{len(used_domains)}"
        
        if domain not in used_domains:
            used_domains[domain] = 0
        used_domains[domain] += 1
        
        if used_domains[domain] > 1:
            return f"comp_{domain}_{used_domains[domain]}"
        return f"comp_{domain}"
    except Exception as e:
        return f"error_domain_{len(used_domains)}"

def validate_urls(urls: List[str]) -> tuple[bool, List[str]]:
    """Validate list of URLs and return validation status and error messages."""
    errors = []
    valid_urls = []
    
    for url in urls:
        if not url or not url.strip():
            errors.append("Empty URL provided")
            continue
            
        url = url.strip()
        
        if not validators.url(url):
            errors.append(f"Invalid URL format: {url}")
            continue
            
        # Check if URL is accessible (basic check)
        try:
            parsed = urlparse(url)
            if not parsed.scheme or not parsed.netloc:
                errors.append(f"URL missing scheme or domain: {url}")
                continue
        except Exception as e:
            errors.append(f"URL parsing error for {url}: {str(e)}")
            continue
        
        valid_urls.append(url)
    
    return len(errors) == 0, errors

def create_excel_report(analysis_results: Dict[str, Any]) -> bytes:
    """Create Excel report from analysis results with enhanced formatting."""
    output = io.BytesIO()
    workbook = Workbook()
    
    # Remove the default sheet
    if "Sheet" in workbook.sheetnames:
        del workbook["Sheet"]
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    used_domains = {}
    competitor_names = {}
    if "competitor_urls" in analysis_results:
        for url in analysis_results["competitor_urls"]:
            competitor_names[url] = get_competitor_name(url, used_domains)

    # --- Summary Sheet ---
    summary_sheet = workbook.create_sheet("Summary")
    summary_data = [
        ["Analysis Summary", ""],
        ["Client URL", analysis_results.get("client_url", "")],
        ["Number of Competitors", len(analysis_results.get("competitor_urls", []))],
        ["Analysis Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Total Entities Found", len(analysis_results.get("comparison_results", {}).get("missing_entities", {}))],
        ["Total Keywords Found", len(analysis_results.get("comparison_results", {}).get("missing_keywords", {}))],
    ]
    
    for row in summary_data:
        summary_sheet.append(row)
    
    # Style the summary sheet
    summary_sheet['A1'].font = header_font
    summary_sheet['A1'].fill = header_fill
    adjust_column_width(summary_sheet, ["Category", "Value"])

    # --- Entity Analysis Sheet ---
    entity_sheet = workbook.create_sheet("Entity Analysis")
    entity_headers = ["Source", "Entity", "Type", "Salience", "Sentiment Score", "Sentiment Magnitude", "Mentions"]
    entity_sheet.append(entity_headers)
    
    # Apply header styling
    for col_num, header in enumerate(entity_headers, 1):
        cell = entity_sheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    entity_data = []
    if "analysis_results" in analysis_results:
        client_analysis = analysis_results["analysis_results"][0]
        for entity_name, data in client_analysis.get("entities", {}).items():
            entity_data.append([
                "Client Page - " + analysis_results.get("client_url", ""),
                entity_name,
                data.get("type", ""),
                data.get("salience", 0),
                data.get("sentiment", {}).get("score", 0),
                data.get("sentiment", {}).get("magnitude", 0),
                ", ".join([mention.get("text", "") for mention in data.get("mentions", [])])
            ])
        
        if len(analysis_results["analysis_results"]) > 1:
            for i, analysis in enumerate(analysis_results["analysis_results"][1:]):
                url = analysis_results["competitor_urls"][i]
                for entity_name, data in analysis.get("entities", {}).items():
                    entity_data.append([
                        f"Competitor - {competitor_names.get(url, 'Unknown')}",
                        entity_name,
                        data.get("type", ""),
                        data.get("salience", 0),
                        data.get("sentiment", {}).get("score", 0),
                        data.get("sentiment", {}).get("magnitude", 0),
                        ", ".join([mention.get("text", "") for mention in data.get("mentions", [])])
                    ])
    
    for row in entity_data:
        entity_sheet.append(row)
    adjust_column_width(entity_sheet, entity_headers)
    
    # --- Keyword Analysis Sheet ---
    keyword_sheet = workbook.create_sheet("Keyword Analysis")
    keyword_headers = ["Source", "Keyword", "Density", "Count", "TF-IDF", "Phrase Count"]
    keyword_sheet.append(keyword_headers)
    
    # Apply header styling
    for col_num, header in enumerate(keyword_headers, 1):
        cell = keyword_sheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    keyword_data = []
    if "analysis_results" in analysis_results:
        client_analysis = analysis_results["analysis_results"][0]
        for keyword, data in client_analysis.get("keyword_analysis", {}).items():
            keyword_data.append([
                "Client Page - " + analysis_results.get("client_url", ""),
                keyword,
                data.get("density", 0),
                data.get("count", 0),
                data.get("tf_idf", 0),
                data.get("phrase_counts", 0)
            ])
        
        if len(analysis_results["analysis_results"]) > 1:
            for i, analysis in enumerate(analysis_results["analysis_results"][1:]):
                url = analysis_results["competitor_urls"][i]
                for keyword, data in analysis.get("keyword_analysis", {}).items():
                    keyword_data.append([
                        f"Competitor - {competitor_names.get(url, 'Unknown')}",
                        keyword,
                        data.get("density", 0),
                        data.get("count", 0),
                        data.get("tf_idf", 0),
                        data.get("phrase_counts", 0)
                    ])
    
    for row in keyword_data:
        keyword_sheet.append(row)
    adjust_column_width(keyword_sheet, keyword_headers)
    
    # --- Missing Entities Sheet ---
    missing_entities_sheet = workbook.create_sheet("Missing Entities")
    missing_entities_headers = ["Entity", "Type", "Competitor Count", "Max Salience"]
    missing_entities_sheet.append(missing_entities_headers)
    
    # Apply header styling
    for col_num, header in enumerate(missing_entities_headers, 1):
        cell = missing_entities_sheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    missing_entities = []
    if "comparison_results" in analysis_results and "missing_entities" in analysis_results["comparison_results"]:
        for entity_name, data in analysis_results["comparison_results"]["missing_entities"].items():
            competitors = data.get("competitors", {})
            max_salience = max([comp.get("salience", 0) for comp in competitors.values()]) if competitors else 0
            missing_entities.append([
                entity_name,
                data.get("type", ""),
                len(competitors),
                max_salience
            ])
    
    for row in missing_entities:
        missing_entities_sheet.append(row)
    adjust_column_width(missing_entities_sheet, missing_entities_headers)
    
    # --- Document Sentiment Sheet ---
    sentiment_sheet = workbook.create_sheet("Document Sentiment")
    sentiment_headers = ["Source", "Score", "Magnitude"]
    sentiment_sheet.append(sentiment_headers)
    
    # Apply header styling
    for col_num, header in enumerate(sentiment_headers, 1):
        cell = sentiment_sheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    sentiment_data = []
    if "analysis_results" in analysis_results:
        client_analysis = analysis_results["analysis_results"][0]
        sentiment_data.append([
            "Client Page - " + analysis_results.get("client_url", ""),
            client_analysis.get("document_sentiment", {}).get("score", 0),
            client_analysis.get("document_sentiment", {}).get("magnitude", 0)
        ])
        
        if len(analysis_results["analysis_results"]) > 1:
            for i, analysis in enumerate(analysis_results["analysis_results"][1:]):
                url = analysis_results["competitor_urls"][i]
                sentiment_data.append([
                    f"Competitor - {competitor_names.get(url, 'Unknown')}",
                    analysis.get("document_sentiment", {}).get("score", 0),
                    analysis.get("document_sentiment", {}).get("magnitude", 0)
                ])
    
    for row in sentiment_data:
        sentiment_sheet.append(row)
    adjust_column_width(sentiment_sheet, sentiment_headers)
    
    workbook.save(output)
    output.seek(0)
    return output.getvalue()

def adjust_column_width(sheet, headers):
    """Adjust column widths based on content."""
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        max_length = len(str(header))
        
        for cell in sheet[column_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        adjusted_width = min(max_length + 2, 75)  # Cap at 75 characters
        sheet.column_dimensions[column_letter].width = adjusted_width

def create_visualization_data(analysis_results: Dict[str, Any]) -> Dict[str, Any]:
    """Create data structures for visualizations."""
    try:
        vis_data = {}
        
        # Entity salience comparison
        if "analysis_results" in analysis_results:
            client_entities = analysis_results["analysis_results"][0].get("entities", {})
            entity_names = list(client_entities.keys())[:10]  # Top 10 entities
            entity_saliences = [client_entities[name]["salience"] for name in entity_names]
            
            vis_data["entity_salience"] = {
                "entities": entity_names,
                "saliences": entity_saliences
            }
        
        # Missing entities by competitor
        if "comparison_results" in analysis_results:
            missing_entities = analysis_results["comparison_results"].get("missing_entities", {})
            competitor_counts = {}
            
            for entity_name, data in missing_entities.items():
                competitors = data.get("competitors", {})
                for comp_name in competitors.keys():
                    competitor_counts[comp_name] = competitor_counts.get(comp_name, 0) + 1
            
            vis_data["competitor_missing_entities"] = competitor_counts
        
        # Sentiment analysis
        if "analysis_results" in analysis_results:
            sentiments = []
            sources = []
            
            for i, analysis in enumerate(analysis_results["analysis_results"]):
                if i == 0:
                    sources.append("Client")
                else:
                    sources.append(f"Competitor {i}")
                
                sentiment = analysis.get("document_sentiment", {})
                sentiments.append(sentiment.get("score", 0))
            
            vis_data["sentiment_comparison"] = {
                "sources": sources,
                "sentiments": sentiments
            }
        
        return vis_data
    
    except Exception as e:
        st.error(f"Error creating visualization data: {e}")
        return {}

def export_to_json(data: Dict[str, Any], filename: str = "analysis_results.json") -> str:
    """Export analysis results to JSON format."""
    try:
        # Create a temporary file
        with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as f:
            json.dump(data, f, indent=2, default=str)
            return f.name
    except Exception as e:
        st.error(f"Error exporting to JSON: {e}")
        return None

def create_pdf_report(analysis_results: Dict[str, Any]) -> bytes:
    """Create a PDF report from analysis results."""
    # This would require additional dependencies like reportlab
    # For now, return the markdown content as bytes
    try:
        if "final_state" in analysis_results:
            markdown_content = analysis_results["final_state"].to_markdown
            return markdown_content.encode('utf-8')
        return b"PDF generation not implemented"
    except Exception as e:
        return f"Error generating PDF: {e}".encode('utf-8')

# Streamlit UI helper functions
def display_analysis_summary(analysis_results: Dict[str, Any]):
    """Display analysis summary in Streamlit."""
    st.subheader("📊 Analysis Summary")
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric(
            "Competitors Analyzed",
            len(analysis_results.get("competitor_urls", []))
        )
    
    with col2:
        missing_entities = analysis_results.get("comparison_results", {}).get("missing_entities", {})
        st.metric(
            "Missing Entities Found",
            len(missing_entities)
        )
    
    with col3:
        missing_keywords = analysis_results.get("comparison_results", {}).get("missing_keywords", {})
        st.metric(
            "Missing Keywords Found",
            len(missing_keywords)
        )
    
    with col4:
        if "analysis_results" in analysis_results and analysis_results["analysis_results"]:
            client_sentiment = analysis_results["analysis_results"][0].get("document_sentiment", {})
            sentiment_score = client_sentiment.get("score", 0)
            st.metric(
                "Client Sentiment Score",
                f"{sentiment_score:.2f}"
            )

def display_entity_visualizations(analysis_results: Dict[str, Any]):
    """Display entity analysis visualizations."""
    st.subheader("📈 Entity Analysis Visualizations")
    
    vis_data = create_visualization_data(analysis_results)
    
    if "entity_salience" in vis_data:
        st.subheader("Top Client Entities by Salience")
        
        fig = px.bar(
            x=vis_data["entity_salience"]["entities"],
            y=vis_data["entity_salience"]["saliences"],
            title="Entity Salience Scores",
            labels={"x": "Entities", "y": "Salience Score"}
        )
        fig.update_layout(height=400)
        st.plotly_chart(fig, use_container_width=True)
    
    if "competitor_missing_entities" in vis_data:
        st.subheader("Missing Entities by Competitor")
        
        competitor_data = vis_data["competitor_missing_entities"]
        fig = px.pie(
            values=list(competitor_data.values()),
            names=list(competitor_data.keys()),
            title="Distribution of Missing Entities by Competitor"
        )
        st.plotly_chart(fig, use_container_width=True)
    
    if "sentiment_comparison" in vis_data:
        st.subheader("Sentiment Analysis Comparison")
        
        sentiment_data = vis_data["sentiment_comparison"]
        fig = px.bar(
            x=sentiment_data["sources"],
            y=sentiment_data["sentiments"],
            title="Document Sentiment Comparison",
            labels={"x": "Sources", "y": "Sentiment Score"},
            color=sentiment_data["sentiments"],
            color_continuous_scale="RdYlBu"
        )
        fig.update_layout(height=400)
        st.plotly_chart(fig, use_container_width=True)
